from __future__ import annotations

import calendar
import csv
import io
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import BinaryIO

from django.db import transaction
from openpyxl import load_workbook

from navapp.models import NAVRecord, ShareClass, month_end


class ImportValidationError(ValueError):
    pass


@dataclass(slots=True)
class ImportRow:
    valuation_month: date
    valuation_date: date
    nav_per_share: Decimal
    status: str = NAVRecord.Status.OFFICIAL
    note: str = ""
    raw_source_date: str = ""
    source_sheet: str = ""
    source_cell: str = ""
    warnings: list[str] = field(default_factory=list)

    def serializable(self) -> dict[str, object]:
        data = asdict(self)
        data["valuation_month"] = self.valuation_month.isoformat()
        data["valuation_date"] = self.valuation_date.isoformat()
        data["nav_per_share"] = str(self.nav_per_share)
        return data


def parse_date(value: object, field_name: str) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m", "%Y/%m"):
        try:
            parsed = datetime.strptime(text, fmt).date()
            return parsed
        except ValueError:
            continue
    raise ImportValidationError(f"{field_name} 無效：{text!r}")


def normalize_row(data: dict[str, object], row_number: int) -> ImportRow:
    try:
        raw_month = parse_date(data.get("valuation_month"), "valuation_month")
        raw_date = parse_date(data.get("valuation_date") or raw_month, "valuation_date")
        nav = Decimal(str(data.get("nav_per_share", "")).strip())
    except (InvalidOperation, ImportValidationError) as exc:
        raise ImportValidationError(f"第 {row_number} 列：{exc}") from exc
    if nav <= 0:
        raise ImportValidationError(f"第 {row_number} 列：每股 NAV 必須大於零。")
    status = str(data.get("status") or NAVRecord.Status.OFFICIAL).strip().upper()
    if status not in NAVRecord.Status.values:
        raise ImportValidationError(f"第 {row_number} 列：狀態 {status!r} 無效。")
    return ImportRow(
        valuation_month=month_end(raw_month),
        valuation_date=raw_date,
        nav_per_share=nav,
        status=status,
        note=str(data.get("note") or "").strip(),
        raw_source_date=str(data.get("raw_source_date") or raw_date.isoformat()),
        source_sheet=str(data.get("source_sheet") or "bulk import"),
        source_cell=str(data.get("source_cell") or f"row {row_number}"),
    )


def validate_sequence(rows: list[ImportRow]) -> None:
    seen: set[date] = set()
    for row in sorted(rows, key=lambda item: item.valuation_month):
        if row.valuation_month in seen:
            raise ImportValidationError(f"匯入資料有重複月份：{row.valuation_month:%Y-%m}。")
        seen.add(row.valuation_month)
    ordered = sorted(seen)
    for previous, current in zip(ordered, ordered[1:], strict=False):
        next_year = previous.year + (previous.month == 12)
        next_month = 1 if previous.month == 12 else previous.month + 1
        expected = date(
            next_year,
            next_month,
            calendar.monthrange(next_year, next_month)[1],
        )
        if current != expected:
            raise ImportValidationError(f"匯入資料缺少月份：{expected:%Y-%m}。")


def parse_csv(file_obj: BinaryIO) -> list[ImportRow]:
    raw = file_obj.read()
    text = raw.decode("utf-8-sig") if isinstance(raw, bytes) else raw
    reader = csv.DictReader(io.StringIO(text))
    required = {"valuation_month", "nav_per_share"}
    if not reader.fieldnames or not required.issubset(reader.fieldnames):
        raise ImportValidationError("CSV 必須包含 valuation_month 及 nav_per_share 欄位標題。")
    rows = [normalize_row(dict(item), index) for index, item in enumerate(reader, start=2)]
    validate_sequence(rows)
    return rows


def parse_xlsx(file_obj: BinaryIO) -> list[ImportRow]:
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    sheet = workbook.active
    values = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "").strip() for value in next(values)]
    except StopIteration as exc:
        raise ImportValidationError("工作簿是空白的。") from exc
    required = {"valuation_month", "nav_per_share"}
    if not required.issubset(headers):
        raise ImportValidationError(
            "XLSX 第一列必須包含 valuation_month 及 nav_per_share 欄位標題。"
        )
    rows: list[ImportRow] = []
    for index, values_row in enumerate(values, start=2):
        if not any(value is not None and str(value).strip() for value in values_row):
            continue
        data = dict(zip(headers, values_row, strict=False))
        data["source_sheet"] = sheet.title
        rows.append(normalize_row(data, index))
    validate_sequence(rows)
    return rows


def parse_uploaded_nav(file_obj, filename: str) -> list[ImportRow]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return parse_csv(file_obj)
    if suffix == ".xlsx":
        return parse_xlsx(file_obj)
    raise ImportValidationError("只支援 .csv 及 .xlsx 檔案。")


def parse_legacy_xsq(path: str | Path, sheet_name: str = "2026 Mar (monthly)") -> list[ImportRow]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ImportValidationError(f"工作表 {sheet_name!r} 不存在。")
    sheet = workbook[sheet_name]
    rows: list[ImportRow] = []
    for row_number in range(3, sheet.max_row + 1):
        raw_date = sheet.cell(row_number, 4).value
        nav = sheet.cell(row_number, 5).value
        if not isinstance(raw_date, (date, datetime)) or nav in {None, ""}:
            continue
        raw_date_value = raw_date.date() if isinstance(raw_date, datetime) else raw_date
        try:
            nav_decimal = Decimal(str(nav))
        except InvalidOperation:
            continue
        if nav_decimal <= 0:
            continue
        warnings: list[str] = []
        normalized_month = month_end(raw_date_value)
        if raw_date_value != normalized_month:
            warnings.append(f"工作簿原始日期 {raw_date_value} 已正規化為 {normalized_month}。")
        if not rows:
            warnings.append(
                "首筆 NAV 與成立日期相同，但數值不同於成立時 NAV；請確認它代表首個每月觀察值。"
            )
        rows.append(
            ImportRow(
                valuation_month=normalized_month,
                valuation_date=normalized_month,
                nav_per_share=nav_decimal,
                raw_source_date=raw_date_value.isoformat(),
                source_sheet=sheet.title,
                source_cell=f"E{row_number}",
                warnings=warnings,
            )
        )
    validate_sequence(rows)
    if not rows:
        raise ImportValidationError("工作簿中找不到具權威性的 NAV 資料列。")
    return rows


@transaction.atomic
def import_rows(
    *,
    share_class: ShareClass,
    rows: list[ImportRow],
    user=None,
    commit: bool = False,
    acknowledge_first_period: bool = False,
) -> dict[str, object]:
    if commit and rows and rows[0].warnings and not acknowledge_first_period:
        raise ImportValidationError("首期對應存在歧義，必須使用 --confirm-first-period 確認。")
    created = 0
    skipped = 0
    conflicts: list[str] = []
    for row in rows:
        existing = NAVRecord.objects.filter(
            share_class=share_class,
            valuation_month=row.valuation_month,
            is_active=True,
        ).first()
        if existing:
            # SQLite's Decimal adapter retains about 15 significant digits, while
            # PostgreSQL NUMERIC retains the full model precision. Treat sub-picounit
            # differences as the same authoritative value so local/demo imports stay
            # idempotent without masking a financially meaningful conflict.
            if abs(existing.nav_per_share - row.nav_per_share) <= Decimal("0.000000000001"):
                skipped += 1
            else:
                conflicts.append(
                    f"{row.valuation_month:%Y-%m}: existing {existing.nav_per_share}, "
                    f"import {row.nav_per_share}"
                )
            continue
        if commit:
            item = NAVRecord(
                share_class=share_class,
                valuation_month=row.valuation_month,
                valuation_date=row.valuation_date,
                nav_per_share=row.nav_per_share,
                status=row.status,
                note=row.note,
                raw_source_date=row.raw_source_date,
                source_sheet=row.source_sheet,
                source_cell=row.source_cell,
                change_acknowledged=bool(row.warnings),
                created_by=user,
                updated_by=user,
            )
            item.full_clean()
            item.save()
        created += 1
    if conflicts:
        raise ImportValidationError("匯入資料衝突：" + "; ".join(conflicts))
    return {
        "mode": "commit" if commit else "dry-run",
        "share_class": str(share_class),
        "proposed": len(rows),
        "created": created,
        "skipped": skipped,
        "rows": [row.serializable() for row in rows],
    }
